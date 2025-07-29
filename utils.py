import asyncio
import httpx
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Tuple, Dict, Any
from config import settings

SEM = asyncio.Semaphore(settings.MAX_CONCURRENCY)

BASE_URL = 'https://api.kucoin.com'

async def get_klines(symbol: str, interval: str = '4hour', limit: int = 60) -> pd.DataFrame:
    url = f'{BASE_URL}/api/v1/market/candles'
    params = {'symbol': symbol, 'type': interval, 'startAt': None, 'endAt': None}
    async with SEM:
        async with httpx.AsyncClient() as client:
            r = await client.get(url, params=params)
            r.raise_for_status()
    data = r.json()['data']
    df = pd.DataFrame(data, columns=['ts','open','close','high','low','volume','turnover'])
    df = df.astype(float)
    df['ts'] = pd.to_datetime(df['ts'], unit='s')
    df = df.sort_values('ts')
    return df[['open','high','low','close','volume','ts']]

async def list_usdt_symbols() -> list:
    url = f'{BASE_URL}/api/v1/symbols'
    async with httpx.AsyncClient() as client:
        r = await client.get(url)
        r.raise_for_status()
    data = r.json()['data']
    symbols = [d['symbol'] for d in data if d['quoteCurrency'] == 'USDT' and d['enable']]
    return symbols

# Indicators

def ema(series: pd.Series, span: int) -> pd.Series:
    return series.ewm(span=span, adjust=False).mean()

def rsi(series: pd.Series, period: int = 14) -> pd.Series:
    delta = series.diff()
    up = delta.clip(lower=0)
    down = -1 * delta.clip(upper=0)
    ma_up = up.rolling(window=period).mean()
    ma_down = down.rolling(window=period).mean()
    rs = ma_up / ma_down
    return 100 - (100 / (1 + rs))

def bollinger_bands(series: pd.Series, window: int = 20, num_std: int = 2) -> Tuple[pd.Series, pd.Series]:
    sma = series.rolling(window=window).mean()
    std = series.rolling(window=window).std()
    upper = sma + num_std * std
    lower = sma - num_std * std
    return upper, lower

# Simple Excel logger

def append_sale_to_excel(path: Path, row: Dict[str, Any]):
    from openpyxl import Workbook, load_workbook
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(row.keys()))
    ws.append(list(row.values()))
    wb.save(path)
